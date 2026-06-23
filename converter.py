import json
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


APP_FOOTER_TEXT = "© 2025 TrackMan Converter by Tom McIntyre"
ALT_FILL = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

COLUMNS = [
    "Time",
    "Club Speed (Mph)", "Ball Speed (Mph)", "Smash Factor",
    "Carry (Yds)", "Total (Yds)",
    "Impact Height (mm)", "Impact Offset (mm)",
    "Club Path (Deg)", "Face Angle (Deg)", "Face To Path (Deg)",
    "Launch Direction (Deg)", "Attack Angle (Deg)",
    "Dynamic Loft (Deg)", "Launch Angle (Deg)", "Spin Loft (Deg)",
    "Spin Rate (Rpm)", "Spin Axis (Deg)",
    "Curve (Ft)", "Carry Side (Ft)", "Total Side (Ft)",
    "Max Height (Ft)", "Landing Angle (Deg)",
    "Swing Direction (Deg)", "Swing Plane (Deg)", "Swing Radius",
    "DPlane Tilt", "Low Point (In)", "Landing Height", "Hang Time (Sec)",
    "Dynamic Lie (Deg)",
]


def _fmt_time(iso: str) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return iso


def _conv_2decimal(v, factor=1.0):
    """Return real numbers (not strings) so Excel sees them as numeric."""
    if v is None:
        return ""
    try:
        return round(float(v) * factor, 2)
    except Exception:
        return ""
    
def _conv_1decimal(v, factor=1.0):
    """Return real numbers (not strings) so Excel sees them as numeric."""
    if v is None:
        return ""
    try:
        return round(float(v) * factor, 1)
    except Exception:
        return ""

def _conv_0decimal(v, factor=1.0):
    """Return real numbers (not strings) so Excel sees them as numeric."""
    if v is None:
        return ""
    try:
        return round(float(v) * factor, 0)
    except Exception:
        return ""

def convert_measurement_to_row(m: dict) -> dict:
    if not m:
        return {k: "" for k in COLUMNS}
    return {
        "Time": _fmt_time(m.get("Time", "")),
        "Club Speed (Mph)": _conv_2decimal(m.get("ClubSpeed"), 2.23694),
        "Ball Speed (Mph)": _conv_2decimal(m.get("BallSpeed"), 2.23694),
        "Smash Factor": _conv_2decimal(m.get("SmashFactor")),
        "Carry (Yds)": _conv_2decimal(m.get("Carry"), 1.09361),
        "Total (Yds)": _conv_2decimal(m.get("Total"), 1.09361),
        "Impact Height (mm)": _conv_2decimal(m.get("ImpactHeight"), 1000),
        "Impact Offset (mm)": _conv_2decimal(m.get("ImpactOffset"), 1000),
        "Club Path (Deg)": _conv_2decimal(m.get("ClubPath")),
        "Face Angle (Deg)": _conv_2decimal(m.get("FaceAngle")),
        "Face To Path (Deg)": _conv_2decimal(m.get("FaceToPath")),
        "Launch Direction (Deg)": _conv_2decimal(m.get("LaunchDirection")),
        "Attack Angle (Deg)": _conv_2decimal(m.get("AttackAngle")),
        "Dynamic Loft (Deg)": _conv_2decimal(m.get("DynamicLoft")),
        "Launch Angle (Deg)": _conv_2decimal(m.get("LaunchAngle")),
        "Spin Loft (Deg)": _conv_2decimal(m.get("SpinLoft")),
        "Spin Rate (Rpm)": _conv_2decimal(m.get("SpinRate")),
        "Spin Axis (Deg)": _conv_2decimal(m.get("SpinAxis")),
        "Curve (Ft)": _conv_2decimal(m.get("Curve"), 3.28084),
        "Carry Side (Ft)": _conv_2decimal(m.get("CarrySide"), 3.28084),
        "Total Side (Ft)": _conv_2decimal(m.get("TotalSide"), 3.28084),
        "Max Height (Ft)": _conv_2decimal(m.get("MaxHeight"), 3.28084),
        "Landing Angle (Deg)": _conv_2decimal(m.get("LandingAngle")),
        "Swing Direction (Deg)": _conv_2decimal(m.get("SwingDirection")),
        "Swing Plane (Deg)": _conv_2decimal(m.get("SwingPlane")),
        "Swing Radius": _conv_2decimal(m.get("SwingRadius")),
        "DPlane Tilt": _conv_2decimal(m.get("DPlaneTilt")),
        "Low Point (In)": _conv_2decimal(m.get("LowPointDistance"), 39.3701),
        "Landing Height": _conv_2decimal(m.get("LandingHeight")),
        "Hang Time (Sec)": _conv_2decimal(m.get("HangTime")),
        "Dynamic Lie (Deg)": _conv_2decimal(m.get("DynamicLie")),
    }


def style_and_finalize_sheet(ws, header_row_idx: int, n_cols: int, n_rows: int):
    ws.row_dimensions[header_row_idx].height = 70

    for c in range(1, n_cols + 1):
        cell = ws.cell(row=header_row_idx, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = header_row_idx + 1
    data_end = header_row_idx + n_rows

    for r in range(data_start, data_end + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value

            if isinstance(val, str):
                try:
                    cell.value = float(val)
                    val = cell.value
                except ValueError:
                    pass

            if isinstance(val, (int, float)):
                if cell.column == 4:  # Smash Factor
                    cell.number_format = "0.00"
                elif cell.column == 13 or cell.column == 28: 
                    cell.number_format = "0.0"
                else:
                     cell.number_format = "0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = f"A{header_row_idx + 1}"
    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row_idx}:{last_col_letter}{data_end}"

    # Time column (A) — 135 px ÷ 7 px/char ≈ 19.3 character units
    ws.column_dimensions["A"].width = 19.3

    summary_labels = ["Pos Av", "Neg Av", "1 Av", "Spread", "% Pos", "% Neg"]
    summary_start = data_end + 2

    for i, label in enumerate(summary_labels):
        cell = ws.cell(row=summary_start + i, column=1, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    pos_row = summary_start
    neg_row = summary_start + 1
    avg_row = summary_start + 2
    spread_row = summary_start + 3
    pct_pos_row = summary_start + 4
    pct_neg_row = summary_start + 5

    for c in range(2, n_cols + 1):
        col_letter = get_column_letter(c)
        data_range = f"{col_letter}{data_start}:{col_letter}{data_end}"

        formulas = [
            f'=IF(COUNTIF({data_range},">0"),AVERAGEIF({data_range},">0"),"—")',
            f'=IF(COUNTIF({data_range},"<0"),AVERAGEIF({data_range},"<0"),"—")',
            f'=IF(COUNTA({data_range}),AVERAGE({data_range}),"—")',
            f'=IF(AND(ISNUMBER({col_letter}{pos_row}),ISNUMBER({col_letter}{neg_row})),'
            f'{col_letter}{pos_row}-{col_letter}{neg_row},"—")',
            f'=IF(COUNTA({data_range}),COUNTIF({data_range},">0")/COUNTA({data_range}),"—")',
            f'=IF(COUNTA({data_range}),COUNTIF({data_range},"<0")/COUNTA({data_range}),"—")',
        ]

        for i, formula in enumerate(formulas):
            row_idx = summary_start + i
            cell = ws.cell(row=row_idx, column=c, value=formula)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if i < 4:
                if cell.column == 4:
                    cell.number_format = "0.00"
                elif cell.column == 13 or cell.column == 28: 
                    cell.number_format = "0.0"
                else:
                     cell.number_format = "0"
            else:
                cell.number_format = "0%"


def append_best_swings(ws, df: pd.DataFrame):
    if df.empty:
        return

    df_num = df.copy()
    for col in df_num.columns:
        if col != "Time":
            df_num[col] = pd.to_numeric(df_num[col], errors="coerce")

    def qualifies(row):
        try:
            return (
                row["Smash Factor"] >= 1.45 and
                abs(row["Impact Height (mm)"]) <= 10 and
                abs(row["Impact Offset (mm)"]) <= 10 and
                abs(row["Club Path (Deg)"]) <= 4 and
                abs(row["Face Angle (Deg)"]) <= 2
            )
        except Exception:
            return False

    q_df = df_num[df_num.apply(qualifies, axis=1)]
    if q_df.empty:
        return

    start_row = ws.max_row + 2

    title_cell = ws.cell(row=start_row, column=1, value="Best Swings")
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    start_row += 1
    first_q_excel_row = start_row

    index_list = list(q_df.index)
    for _, row in q_df.iterrows():
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=start_row, column=col_idx, value=val)
            if isinstance(val, (int, float)):
                if cell.column == 4:
                    cell.number_format = "0.00"
                elif cell.column == 13 or cell.column == 28: 
                    cell.number_format = "0.0"
                else:
                     cell.number_format = "0"
                
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        start_row += 1

    metric_cols = ["Impact Height (mm)", "Impact Offset (mm)",
                   "Club Path (Deg)", "Face Angle (Deg)"]
    try:
        dist = q_df[metric_cols].abs().sum(axis=1)
        best_idx = dist.idxmin()
        offset = index_list.index(best_idx)
        best_excel_row = first_q_excel_row + offset
    except Exception:
        return

    blue_border = Border(
        left=Side(style="thin", color="0000FF"),
        right=Side(style="thin", color="0000FF"),
        top=Side(style="thin", color="0000FF"),
        bottom=Side(style="thin", color="0000FF"),
    )
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=best_excel_row, column=col_idx).border = blue_border


# Labels written by style_and_finalize_sheet / append_best_swings that mark
# the end of real data rows in an existing master sheet.
_MASTER_END_LABELS = {
    "Pos Av", "Neg Av", "1 Av", "Spread", "% Pos", "% Neg", "Best Swings"
}


def append_to_master_workbook(data: dict, master_path: Path) -> None:
    """Append session stroke data into the master workbook at *master_path*.

    Each club (by name only, tags ignored) gets its own sheet.  If the file
    does not exist it is created from scratch.  When a club sheet already
    exists the existing data rows are read back, the new rows are appended
    below them, and the whole sheet is rebuilt so the summary statistics and
    Best-Swings section reflect the full accumulated dataset.
    """
    if master_path.exists():
        wb = load_workbook(str(master_path))
    else:
        wb = Workbook()
        wb.remove(wb.active)

    stroke_groups = data.get("StrokeGroups", []) or []

    for g in stroke_groups:
        club = str(g.get("Club", "Unknown Club"))
        sheet_name = club[:31]

        # Collect new rows from this session group.
        new_rows = []
        for s in g.get("Strokes", []) or []:
            m = s.get("Measurement")
            if isinstance(m, dict):
                new_rows.append(convert_measurement_to_row(m))

        if not new_rows:
            continue

        # Read back existing data rows (skip header, stop at summary label or
        # blank separator row).
        existing_rows = []
        existing_times: set[str] = set()
        sheet_idx = None
        if sheet_name in wb.sheetnames:
            sheet_idx = wb.sheetnames.index(sheet_name)
            ws_old = wb[sheet_name]
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                first = row[0] if row else None
                if first is None or first in _MASTER_END_LABELS:
                    break
                row_dict = {col: (val if val is not None else "") for col, val in zip(COLUMNS, row)}
                existing_rows.append(row_dict)
                # Normalise the Time value to a plain string for dedup comparisons.
                # openpyxl returns stored datetimes as Python datetime objects.
                t = row_dict.get("Time", "")
                if t:
                    t_str = t.strftime("%Y-%m-%d %H:%M:%S") if hasattr(t, "strftime") else str(t)
                    existing_times.add(t_str)
            del wb[sheet_name]

        # Filter new rows – skip any stroke whose Time already appears in the master.
        new_rows = [row for row in new_rows if row.get("Time", "") not in existing_times]
        if not new_rows:
            # Nothing genuinely new for this club; restore sheet as-is.
            if existing_rows:
                ws = wb.create_sheet(title=sheet_name, index=sheet_idx)
                df_restore = pd.DataFrame(existing_rows, columns=COLUMNS)
                for col in df_restore.columns:
                    if col == "Time":
                        df_restore[col] = pd.to_datetime(df_restore[col], errors="coerce")
                    else:
                        df_restore[col] = pd.to_numeric(df_restore[col], errors="coerce")
                for r in dataframe_to_rows(df_restore, index=False, header=True):
                    ws.append(r)
                style_and_finalize_sheet(ws, 1, len(COLUMNS), len(df_restore))
                append_best_swings(ws, df_restore)
            continue

        # Rebuild sheet at the same position (or appended when new).
        ws = wb.create_sheet(title=sheet_name, index=sheet_idx)

        all_rows = existing_rows + new_rows
        df = pd.DataFrame(all_rows, columns=COLUMNS)
        for col in df.columns:
            if col == "Time":
                df[col] = pd.to_datetime(df[col], errors="coerce")
            else:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        style_and_finalize_sheet(ws, 1, len(COLUMNS), len(df))
        append_best_swings(ws, df)

    wb.save(str(master_path))


def split_by_player(data: dict) -> dict[str, dict]:
    """Partition a report's StrokeGroups by player.

    Returns a dict mapping each player's display name to a copy of *data*
    containing only that player's StrokeGroups.  When no Player field is
    present on any group the whole report is returned under the key
    "Unknown Player".
    """
    groups_by_player: dict[str, list] = {}
    for g in data.get("StrokeGroups", []) or []:
        player_info = g.get("Player") or {}
        name = (player_info.get("Name") or "").strip() or "Unknown Player"
        groups_by_player.setdefault(name, []).append(g)

    result: dict[str, dict] = {}
    for name, groups in groups_by_player.items():
        player_data = dict(data)          # shallow copy preserves top-level keys
        player_data["StrokeGroups"] = groups
        result[name] = player_data
    return result


def _safe_filename_part(name: str) -> str:
    """Convert a player name to a safe filename fragment (no special chars)."""
    import re as _re
    return _re.sub(r'[^A-Za-z0-9_-]+', '_', name).strip('_')


def build_workbook_per_club(data: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    stroke_groups = data.get("StrokeGroups", []) or []
    all_rows = []
    used_titles: dict[str, int] = {}

    for g in stroke_groups:
        club = str(g.get("Club", "Unknown Club"))

        # Build tag suffix from the StrokeGroup's Tags array.
        # Tags may be plain strings or objects with a "Name" key.
        raw_tags = g.get("Tags") or []
        tag_strs = []
        for t in raw_tags:
            if isinstance(t, str):
                tag_strs.append(t)
            elif isinstance(t, dict):
                label = t.get("Name") or t.get("Label") or ""
                if label:
                    tag_strs.append(label)
        base_title = f"{club} - {', '.join(tag_strs)}" if tag_strs else club

        # Excel sheet names cannot contain: / \ * ? [ ] :
        for _ch in r'/\*?[]':
            base_title = base_title.replace(_ch, "-")
        base_title = base_title.replace(":", "-")

        # Excel sheet names are limited to 31 characters; deduplicate within workbook.
        base_title = base_title[:31]
        if base_title in used_titles:
            used_titles[base_title] += 1
            suffix = f" ({used_titles[base_title]})"
            title = base_title[: 31 - len(suffix)] + suffix
        else:
            used_titles[base_title] = 1
            title = base_title

        ws = wb.create_sheet(title=title)

        rows = []
        for s in g.get("Strokes", []) or []:
            m = s.get("Measurement")
            if isinstance(m, dict):
                rows.append(convert_measurement_to_row(m))

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=COLUMNS)
        for col in df.columns:
            # Treat the Time column as datetimes; all other columns numeric.
            if col == "Time":
                df[col] = pd.to_datetime(df[col], format="%Y-%m-%d %H:%M:%S", errors='coerce')
            else:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        style_and_finalize_sheet(ws, 1, len(COLUMNS), len(df.index))
        append_best_swings(ws, df)

        all_rows.extend(rows)

    if all_rows:
        ws_all = wb.create_sheet("All Data")
        df_all = pd.DataFrame(all_rows)
        for col in df_all.columns:
            # Treat the Time column as datetimes; all other columns numeric.
            if col == "Time":
                df_all[col] = pd.to_datetime(df_all[col], format="%Y-%m-%d %H:%M:%S", errors='coerce')
            else:
                df_all[col] = pd.to_numeric(df_all[col], errors='coerce')

        for r in dataframe_to_rows(df_all, index=False, header=True):
            ws_all.append(r)

        style_and_finalize_sheet(ws_all, 1, len(df_all.columns), len(df_all.index))
    else:
        ws = wb.create_sheet("Trackman Report")
        ws.cell(1, 1, "No StrokeGroups found in the JSON.")

    return wb

"""
Unified TrackMan Converter GUI Application
Combines all GUI dialogs, conversion logic, and Excel workbook building in a single module.
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
import json
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import trackman_auth
from trackman_api import download_report, get_all_report_ids_from_chrome, fetch_report_metadata_batch


# ============================================================================
# Constants & Configuration
# ============================================================================

APP_FOOTER_TEXT = "© 2025 TrackMan Converter by Tom McIntyre"
TRACKMAN_COLOUR = "#001AFF"
DARK_BG = "#1E1E1E"

COLUMNS = [
    "Time",
    "Club Speed (Mph)", "Ball Speed (Mph)", "Smash Factor",
    "Carry (Yds)", "Total (Yds)",
    "Impact Height (mm)", "Impact Offset (mm)",
    "Club Path (Deg)", "Face Angle (Deg)", "Face To Path (Deg)",
    "Launch Direction (Deg)", "Attack Angle (Deg)",
    "Dynamic Loft (Deg)", "Launch Angle (Deg)", "Spin Loft (Deg)",
    "Spin Rate (Rpm)", "Spin Axis (Deg)",
    "Curve (Ft)", "Carry Side (Ft)", "Total Side (Ft)",
    "Max Height (Ft)", "Landing Angle (Deg)",
    "Swing Direction (Deg)", "Swing Plane (Deg)", "Swing Radius",
    "DPlane Tilt", "Low Point (In)", "Landing Height", "Hang Time (Sec)",
    "Dynamic Lie (Deg)",
]

ALT_FILL = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")


# ============================================================================
# Data Conversion Utilities
# ============================================================================

def _fmt_time(iso: str) -> str:
    """Format ISO timestamp to readable string."""
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return iso


def _conv(v, factor=1.0):
    """Convert value to float with optional scaling factor.
    Returns real numbers (not strings) so Excel sees them as numeric.
    """
    if v is None:
        return ""
    try:
        return round(float(v) * factor, 2)
    except Exception:
        return ""


def convert_measurement_to_row(m: dict) -> dict:
    """Convert a measurement dictionary to an Excel row dictionary."""
    if not m:
        return {k: "" for k in COLUMNS}
    return {
        "Time": _fmt_time(m.get("Time", "")),
        "Club Speed (Mph)": _conv(m.get("ClubSpeed"), 2.23694),
        "Ball Speed (Mph)": _conv(m.get("BallSpeed"), 2.23694),
        "Smash Factor": _conv(m.get("SmashFactor")),
        "Carry (Yds)": _conv(m.get("Carry"), 1.09361),
        "Total (Yds)": _conv(m.get("Total"), 1.09361),
        "Impact Height (mm)": _conv(m.get("ImpactHeight"), 1000),
        "Impact Offset (mm)": _conv(m.get("ImpactOffset"), 1000),
        "Club Path (Deg)": _conv(m.get("ClubPath")),
        "Face Angle (Deg)": _conv(m.get("FaceAngle")),
        "Face To Path (Deg)": _conv(m.get("FaceToPath")),
        "Launch Direction (Deg)": _conv(m.get("LaunchDirection")),
        "Attack Angle (Deg)": _conv(m.get("AttackAngle")),
        "Dynamic Loft (Deg)": _conv(m.get("DynamicLoft")),
        "Launch Angle (Deg)": _conv(m.get("LaunchAngle")),
        "Spin Loft (Deg)": _conv(m.get("SpinLoft")),
        "Spin Rate (Rpm)": _conv(m.get("SpinRate")),
        "Spin Axis (Deg)": _conv(m.get("SpinAxis")),
        "Curve (Ft)": _conv(m.get("Curve"), 3.28084),
        "Carry Side (Ft)": _conv(m.get("CarrySide"), 3.28084),
        "Total Side (Ft)": _conv(m.get("TotalSide"), 3.28084),
        "Max Height (Ft)": _conv(m.get("MaxHeight"), 3.28084),
        "Landing Angle (Deg)": _conv(m.get("LandingAngle")),
        "Swing Direction (Deg)": _conv(m.get("SwingDirection")),
        "Swing Plane (Deg)": _conv(m.get("SwingPlane")),
        "Swing Radius": _conv(m.get("SwingRadius")),
        "DPlane Tilt": _conv(m.get("DPlaneTilt")),
        "Low Point (In)": _conv(m.get("LowPointDistance"), 39.3701),
        "Landing Height": _conv(m.get("LandingHeight")),
        "Hang Time (Sec)": _conv(m.get("HangTime")),
        "Dynamic Lie (Deg)": _conv(m.get("DynamicLie")),
    }


def style_and_finalize_sheet(ws, header_row_idx: int, n_cols: int, n_rows: int):
    """Apply formatting to worksheet: headers, alignment, summary statistics, filters."""
    ws.row_dimensions[header_row_idx].height = 70

    for c in range(1, n_cols + 1):
        cell = ws.cell(row=header_row_idx, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = header_row_idx + 1
    data_end = header_row_idx + n_rows

    for r in range(data_start, data_end + 1):
        fill = ALT_FILL if (r - data_start) % 2 == 0 else None
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value

            if isinstance(val, str):
                try:
                    cell.value = float(val)
                    val = cell.value
                except ValueError:
                    pass

            if isinstance(val, (int, float)):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if fill:
                cell.fill = ALT_FILL

    ws.freeze_panes = f"A{header_row_idx + 1}"
    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row_idx}:{last_col_letter}{data_end}"

    summary_labels = ["Pos Av", "Neg Av", "1 Av", "Spread", "% Pos", "% Neg"]
    summary_start = data_end + 2

    for i, label in enumerate(summary_labels):
        cell = ws.cell(row=summary_start + i, column=1, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    pos_row = summary_start
    neg_row = summary_start + 1
    avg_row = summary_start + 2
    spread_row = summary_start + 3
    pct_pos_row = summary_start + 4
    pct_neg_row = summary_start + 5

    for c in range(2, n_cols + 1):
        col_letter = get_column_letter(c)
        data_range = f"{col_letter}{data_start}:{col_letter}{data_end}"

        formulas = [
            f'=IF(COUNTIF({data_range},">0"),AVERAGEIF({data_range},">0"),"—")',
            f'=IF(COUNTIF({data_range},"<0"),AVERAGEIF({data_range},"<0"),"—")',
            f'=IF(COUNTA({data_range}),AVERAGE({data_range}),"—")',
            f'=IF(AND(ISNUMBER({col_letter}{pos_row}),ISNUMBER({col_letter}{neg_row})),'
            f'{col_letter}{pos_row}-{col_letter}{neg_row},"—")',
            f'=IF(COUNTA({data_range}),COUNTIF({data_range},">0")/COUNTA({data_range}),"—")',
            f'=IF(COUNTA({data_range}),COUNTIF({data_range},"<0")/COUNTA({data_range}),"—")',
        ]

        for i, formula in enumerate(formulas):
            row_idx = summary_start + i
            cell = ws.cell(row=row_idx, column=c, value=formula)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if i < 4:
                cell.number_format = "0.00"
            else:
                cell.number_format = "0%"


def append_best_swings(ws, df: pd.DataFrame):
    """Append a 'Best Swings' section to the worksheet with best-fit swing highlighted."""
    if df.empty:
        return

    df_num = df.copy()
    for col in df_num.columns:
        df_num[col] = pd.to_numeric(df_num[col], errors="coerce")

    def qualifies(row):
        try:
            return (
                row["Smash Factor"] >= 1.45 and
                abs(row["Impact Height (mm)"]) <= 10 and
                abs(row["Impact Offset (mm)"]) <= 10 and
                abs(row["Club Path (Deg)"]) <= 4 and
                abs(row["Face Angle (Deg)"]) <= 2
            )
        except Exception:
            return False

    q_df = df_num[df_num.apply(qualifies, axis=1)]
    if q_df.empty:
        return

    start_row = ws.max_row + 2

    title_cell = ws.cell(row=start_row, column=1, value="Best Swings")
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    start_row += 1
    first_q_excel_row = start_row

    index_list = list(q_df.index)
    for _, row in q_df.iterrows():
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=start_row, column=col_idx, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        start_row += 1

    metric_cols = ["Impact Height (mm)", "Impact Offset (mm)",
                   "Club Path (Deg)", "Face Angle (Deg)"]
    try:
        dist = q_df[metric_cols].abs().sum(axis=1)
        best_idx = dist.idxmin()
        offset = index_list.index(best_idx)
        best_excel_row = first_q_excel_row + offset
    except Exception:
        return

    blue_border = Border(
        left=Side(style="thin", color="0000FF"),
        right=Side(style="thin", color="0000FF"),
        top=Side(style="thin", color="0000FF"),
        bottom=Side(style="thin", color="0000FF"),
    )
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=best_excel_row, column=col_idx).border = blue_border


def build_workbook_per_club(data: dict) -> Workbook:
    """Build an Excel workbook from TrackMan JSON data, with one sheet per club plus 'All Data'."""
    wb = Workbook()
    wb.remove(wb.active)

    stroke_groups = data.get("StrokeGroups", []) or []
    all_rows = []

    for g in stroke_groups:
        club = str(g.get("Club", "Unknown Club"))
        ws = wb.create_sheet(title=club[:31])

        rows = []
        for s in g.get("Strokes", []) or []:
            m = s.get("Measurement")
            if isinstance(m, dict):
                rows.append(convert_measurement_to_row(m))

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=COLUMNS)
        for col in df.columns:
            # Treat the Time column as datetimes; all other columns numeric.
            if col == "Time":
                df[col] = pd.to_datetime(df[col], format="%Y-%m-%d %H:%M:%S", errors='coerce')
            else:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        style_and_finalize_sheet(ws, 1, len(COLUMNS), len(df.index))
        append_best_swings(ws, df)

        all_rows.extend(rows)

    if all_rows:
        ws_all = wb.create_sheet("All Data")
        df_all = pd.DataFrame(all_rows)
        for col in df_all.columns:
            # Treat the Time column as datetimes; all other columns numeric.
            if col == "Time":
                df_all[col] = pd.to_datetime(df_all[col], format="%Y-%m-%d %H:%M:%S", errors='coerce')
            else:
                df_all[col] = pd.to_numeric(df_all[col], errors='coerce')

        for r in dataframe_to_rows(df_all, index=False, header=True):
            ws_all.append(r)

        style_and_finalize_sheet(ws_all, 1, len(df_all.columns), len(df_all.index))
    else:
        ws = wb.create_sheet("Trackman Report")
        ws.cell(1, 1, "No StrokeGroups found in the JSON.")

    return wb


def convert_json_to_excel(json_path: str, out_path: str = None):
    """Convert a TrackMan JSON to Excel.

    If `out_path` is provided, the workbook is saved there without showing a file dialog.
    If `out_path` is None, the GUI save dialog is shown (original behavior).
    Returns a `Path` to the saved file, or `None` if the user cancelled the dialog.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = build_workbook_per_club(data)

    if out_path:
        wb.save(out_path)
        return Path(out_path)

    default_name = datetime.now().strftime("Trackman_Report_%Y%m%d_%H%M%S.xlsx")
    default_dir = str(Path.home() / "Documents")

    save_path = filedialog.asksaveasfilename(
        title="Save Converted Excel File",
        defaultextension=".xlsx",
        initialfile=default_name,
        initialdir=default_dir,
        filetypes=[("Excel Files", "*.xlsx")],
    )
    if not save_path:
        messagebox.showinfo("Cancelled", "Save cancelled. File not created.")
        return None

    wb.save(save_path)
    return Path(save_path)


# ============================================================================
# GUI Dialog Classes
# ============================================================================

class SimpleLoadingDialog(ctk.CTkToplevel):
    """Simple loading dialog for download/convert operations."""
    
    def __init__(self, parent, title="Processing"):
        super().__init__(parent)
        self.title(title)
        self.geometry("400x150")
        self.grab_set()
        self.configure(fg_color="#1E1E1E")
        self.resizable(False, False)
        
        frame = ctk.CTkFrame(self, fg_color="#1E1E1E")
        frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        self.label = ctk.CTkLabel(
            frame,
            text="Processing...",
            font=("Segoe UI", 16),
            text_color="#AAAAAA",
        )
        self.label.pack(pady=50)
    
    def update_text(self, text):
        """Update the loading dialog text."""
        self.label.configure(text=text)
        self.update_idletasks()


class TokenAndReportDialog(ctk.CTkToplevel):
    """Unified modal dialog: handles token, report fetching, and selection all in one window."""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.title("TrackMan Report Selection")
        self.geometry("720x600")
        self.grab_set()
        self.configure(fg_color="#1E1E1E")
        
        self.token = None
        self.reports = []
        self.selected_report = None
        self.parent_app = parent
        
        # Main frame with title and content area
        self.title_label = ctk.CTkLabel(
            self,
            text="Fetching TrackMan Reports",
            font=("Segoe UI", 22, "bold"),
            text_color="white",
        )
        self.title_label.pack(pady=(20, 5))
        
        # Content frame will hold loading state or report selector
        self.content_frame = ctk.CTkFrame(self, fg_color="#1E1E1E")
        self.content_frame.pack(fill="both", expand=True, padx=40, pady=20)
        
        self._show_loading_prompt()
    
    def _show_loading_prompt(self, text="Fetching TrackMan reports..."):
        """Show a loading message inside the dialog."""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        label = ctk.CTkLabel(
            self.content_frame,
            text=text,
            font=("Segoe UI", 16),
            text_color="#AAAAAA",
        )
        label.pack(pady=100)
        self.update_idletasks()
    
    def _show_report_selector(self):
        """Display the grid of selectable report cards."""
        self.title_label.configure(text="Select a TrackMan Report")
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        scroll_area = ctk.CTkScrollableFrame(self.content_frame, fg_color="#1E1E1E")
        scroll_area.pack(fill="both", expand=True)
        
        container = ctk.CTkFrame(scroll_area, fg_color="#1E1E1E")
        container.pack(anchor="center")
        
        # Sort reports by date (newest first)
        sorted_reports = sorted(
            self.reports, 
            key=lambda r: r.get("time", datetime.utcnow()), 
            reverse=True
        )
        
        cols = 3
        
        for i, r in enumerate(sorted_reports):
            date = r.get("time", datetime.utcnow())
            month = date.strftime("%b").upper()
            day = date.strftime("%d")
            year = date.strftime("%Y")
            
            frame = ctk.CTkFrame(
                container,
                fg_color="#2A2A2A",
                corner_radius=12,
                border_color="#444",
                border_width=1,
                width=180,
                height=160,
            )
            frame.grid(row=i // cols, column=i % cols, padx=18, pady=18)
            
            ctk.CTkLabel(
                frame,
                text=f"{month}\n{day}",
                font=("Segoe UI", 20, "bold"),
                text_color=TRACKMAN_COLOUR,
                justify="center",
            ).pack(pady=(10, 4))
            
            ctk.CTkLabel(
                frame,
                text="Multi Group Report",
                font=("Segoe UI", 13),
                text_color="white",
            ).pack()
            
            ctk.CTkLabel(
                frame,
                text=year,
                font=("Segoe UI", 11, "italic"),
                text_color="#AAAAAA",
            ).pack(pady=(0, 8))
            
            ctk.CTkButton(
                frame,
                text="Select",
                fg_color=TRACKMAN_COLOUR,
                hover_color="#FF8533",
                text_color="white",
                corner_radius=8,
                height=28,
                width=100,
                font=("Segoe UI", 13, "bold"),
                command=lambda rep=r: self._on_select(rep),
            ).pack(pady=(5, 10))
    
    def _on_select(self, report):
        """Set the selected report and close the dialog."""
        self.selected_report = report
        self.destroy()
    
    def populate_reports(self, token, reports):
        """Update dialog with token and reports, then show selector."""
        self.token = token
        self.reports = reports
        self._show_report_selector()


# ============================================================================
# Main Application Class
# ============================================================================

class LoadingOverlay(ctk.CTkToplevel):
    """Full overlay loading screen for the main window."""
    
    def __init__(self, parent, text="Loading..."):
        super().__init__(parent)
        self.attributes("-topmost", True)
        
        # Make the overlay fill the parent window
        parent_x = parent.winfo_x()
        parent_y = parent.winfo_y()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        self.geometry(f"{parent_width}x{parent_height}+{parent_x}+{parent_y}")
        self.resizable(False, False)
        self.overrideredirect(True)
        
        frame = ctk.CTkFrame(self, fg_color=DARK_BG)
        frame.pack(fill="both", expand=True)
        
        self.label = ctk.CTkLabel(
            frame,
            text=text,
            font=("Segoe UI", 18),
            text_color="#AAAAAA",
        )
        self.label.pack(expand=True)
    
    def update_text(self, text):
        """Update the overlay text."""
        self.label.configure(text=text)
        self.update_idletasks()


class TrackmanApp(ctk.CTk):
    """Main TrackMan Converter Application."""
    
    def __init__(self):
        super().__init__()
        self.title("TrackMan Converter")
        self.geometry("700x480")
        self.resizable(False, False)
        self.configure(fg_color=DARK_BG)
        self.overlay = None

        # Header
        header = ctk.CTkFrame(self, fg_color=TRACKMAN_COLOUR, corner_radius=0, height=90)
        header.pack(fill="x")
        ctk.CTkLabel(
            header,
            text="TrackMan Report Converter",
            font=("Segoe UI", 28, "bold"),
            text_color="white",
        ).pack(pady=25)

        # Content
        content = ctk.CTkFrame(self, fg_color=DARK_BG)
        content.pack(expand=True, fill="both")

        ctk.CTkLabel(
            content,
            text="Automatically fetches your latest TrackMan report\n"
                 "and converts it to a formatted Excel file.",
            font=("Segoe UI", 16),
            text_color="#BBBBBB",
            justify="center",
        ).pack(pady=(60, 30))

        ctk.CTkButton(
            content,
            text=" Download & Convert Latest Report",
            width=320,
            height=60,
            corner_radius=12,
            font=("Segoe UI", 18, "bold"),
            fg_color=TRACKMAN_COLOUR,
            hover_color="#FF8533",
            text_color="white",
            command=self.handle_cloud,
        ).pack(pady=10)

        self.status_label = ctk.CTkLabel(
            content, text="", font=("Segoe UI", 14), text_color="#AAAAAA"
        )
        self.status_label.pack(pady=(30, 10))

        # Footer
        ctk.CTkLabel(
            self,
            text=APP_FOOTER_TEXT,
            font=("Segoe UI", 11, "italic"),
            text_color="#666666",
        ).pack(side="bottom", pady=8)

    def show_overlay(self, text="Loading..."):
        """Show the loading overlay."""
        if self.overlay:
            self.overlay.destroy()
        self.overlay = LoadingOverlay(self, text)
        self.overlay.update()

    def hide_overlay(self):
        """Hide the loading overlay."""
        if self.overlay:
            self.overlay.destroy()
            self.overlay = None

    def handle_cloud(self):
        """Orchestrate token retrieval, ChromeHistory search, report selection, download, and conversion."""
        try:
            # Show dialog immediately (in loading state)
            dialog = TokenAndReportDialog(self)
            
            # Fetch token
            dialog._show_loading_prompt("Checking TrackMan login...")
            self.update_idletasks()
            
            token = trackman_auth.get_saved_token() or trackman_auth.login_via_browser()
            if not token:
                dialog.destroy()
                raise Exception("Could not retrieve TrackMan token.")

            # Search Chrome history
            dialog._show_loading_prompt("Searching Chrome history for TrackMan reports...")
            self.update_idletasks()
            
            raw_reports = get_all_report_ids_from_chrome(limit=50)

            if not raw_reports:
                dialog.destroy()
                messagebox.showerror(
                    "No Reports Found",
                    "No recent TrackMan reports were found in Chrome history.\n"
                    "Please open a TrackMan report in Chrome and try again."
                )
                return

            # Deduplicate
            seen = set()
            unique_reports = []
            for r in raw_reports:
                rid = r.get("id")
                if rid and rid not in seen:
                    seen.add(rid)
                    unique_reports.append(r)

            # Fetch metadata
            dialog._show_loading_prompt("Getting upload dates from TrackMan...")
            self.update_idletasks()
            
            report_ids = [r["id"] for r in unique_reports]
            metadata_list = fetch_report_metadata_batch(token, report_ids, max_workers=5)
            
            enriched = []
            for r, meta in zip(unique_reports, metadata_list):
                if meta and meta.get("created"):
                    try:
                        meta["time"] = datetime.fromisoformat(meta["created"].replace("Z", "+00:00"))
                    except Exception:
                        meta["time"] = datetime.utcnow()
                    enriched.append(meta)
                else:
                    enriched.append({"id": r["id"], "time": datetime.utcnow()})

            # Populate dialog with reports and show selector
            dialog.populate_reports(token, enriched)
            
            # Wait for user selection
            self.wait_window(dialog)
            
            if not dialog.selected_report:
                return  # User cancelled
            
            selected_report = dialog.selected_report
            self.show_overlay("Downloading selected report...")
            try:
                json_path = download_report(token, selected_report["id"])
                self.overlay.update_text("Converting to formatted Excel...")
                out_dir = Path(r"C:\Trackman\Data")
                out_dir.mkdir(parents=True, exist_ok=True)
                default_name = f"{selected_report['time'].strftime('%Y_%m_%d')}.xlsx"
                out_path = out_dir / default_name
                result = convert_json_to_excel(json_path, str(out_path))
                self.hide_overlay()
                messagebox.showinfo("Success", f"Downloaded and converted!\nSaved as:\n{result}")
            except Exception as e:
                self.hide_overlay()
                messagebox.showerror("Error", str(e))

        except Exception as e:
            messagebox.showerror("Error", str(e))


# ============================================================================
# Application Entry Point
# ============================================================================

if __name__ == "__main__":
    # Configure UI appearance
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("dark-blue")
    
    app = TrackmanApp()
    app.mainloop()
